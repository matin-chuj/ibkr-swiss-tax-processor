"""
Basel-Landschaft Tax Report Generator
Generates comprehensive tax reports in multiple formats (Excel, PDF, JSON)
"""

import json
import logging
from pathlib import Path
from typing import Dict, List, Optional
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from tax_calculator_bl import TaxCalculatorBL

logger = logging.getLogger(__name__)


class ReportGeneratorBL:
    """Report generator for Basel-Landschaft canton"""

    def __init__(self, parsed_data: Dict, config_file: str = 'basellandschaft_config.json'):
        """
        Initialize report generator
        
        Args:
            parsed_data: Parsed IBKR statement data
            config_file: Path to Basel-Landschaft configuration file
        """
        self.parsed_data = parsed_data
        self.tax_calculator = TaxCalculatorBL(config_file)
        self.config = self.tax_calculator.config
        
        # Generate tax summary
        self.tax_summary = self.tax_calculator.generate_tax_summary(parsed_data)
        
        logger.info("Report generator initialized for Basel-Landschaft")

    def generate_all_reports(self, output_dir: str = '.') -> Dict[str, str]:
        """
        Generate all report formats
        
        Args:
            output_dir: Output directory for reports
            
        Returns:
            Dictionary with paths to generated reports
        """
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)
        
        year = self.tax_summary['tax_year']
        
        reports = {
            'excel': str(output_path / f'Wertschriftenverzeichnis_BL_{year}.xlsx'),
            'pdf': str(output_path / f'Tax_Report_BL_{year}.pdf'),
            'json': str(output_path / f'Tax_Summary_BL_{year}.json')
        }
        
        logger.info("Generating all report formats...")
        
        self.generate_excel_report(reports['excel'])
        self.generate_pdf_report(reports['pdf'])
        self.generate_json_report(reports['json'])
        
        logger.info(f"✅ All reports generated in {output_dir}")
        
        return reports

    def generate_excel_report(self, output_file: str):
        """
        Generate Excel report with Wertschriftenverzeichnis structure
        
        Args:
            output_file: Path to output Excel file
        """
        logger.info(f"📊 Generating Excel report: {output_file}")
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Section 1: Vermögensaufstellung (Asset Position)
            self._write_asset_position_sheet(writer)
            
            # Section 2: Einkünfte aus Vermögen (Income from Assets)
            self._write_income_sheet(writer)
            
            # Section 3: Kapitalgewinne/Verluste (Capital Gains/Losses)
            self._write_capital_gains_sheet(writer)
            
            # Section 4: Kosten und Abzüge (Costs and Deductions)
            self._write_costs_sheet(writer)
            
            # Section 5: Devisenverluste (FX Gains/Losses)
            self._write_fx_sheet(writer)
            
            # Summary sheet
            self._write_summary_sheet(writer)
            
            # Apply formatting
            self._apply_excel_formatting(writer)
        
        logger.info(f"✅ Excel report generated: {output_file}")

    def _write_asset_position_sheet(self, writer):
        """Write Vermögensaufstellung sheet"""
        transactions = self.parsed_data.get('transactions', [])
        open_positions = self.parsed_data.get('open_positions', [])
        
        # Categorize assets
        swiss_stocks = []
        foreign_stocks = []
        cash_chf = 0.0
        foreign_currencies = {}
        
        for pos in open_positions:
            symbol = pos.get('symbol', '')
            value_chf = pos.get('value_chf', 0.0)
            currency = pos.get('currency', 'CHF')
            
            # Simple heuristic for Swiss vs Foreign
            if currency == 'CHF' or len(symbol) <= 4:
                swiss_stocks.append(pos)
            else:
                foreign_stocks.append(pos)
        
        # Build data frame
        data = []
        data.append(['VERMÖGENSAUFSTELLUNG (Stand: 31.12.2025)', ''])
        data.append(['', ''])
        
        # Swiss stocks
        data.append(['Schweizer Aktien', 'Wert (CHF)'])
        total_swiss = 0.0
        for stock in swiss_stocks:
            data.append([stock.get('symbol', ''), f"{stock.get('value_chf', 0.0):.2f}"])
            total_swiss += stock.get('value_chf', 0.0)
        data.append(['Zwischensumme Schweizer Aktien', f"{total_swiss:.2f}"])
        data.append(['', ''])
        
        # Foreign stocks
        data.append(['Ausländische Aktien', 'Wert (CHF)'])
        total_foreign = 0.0
        for stock in foreign_stocks:
            data.append([stock.get('symbol', ''), f"{stock.get('value_chf', 0.0):.2f}"])
            total_foreign += stock.get('value_chf', 0.0)
        data.append(['Zwischensumme Ausländische Aktien', f"{total_foreign:.2f}"])
        data.append(['', ''])
        
        # Totals
        data.append(['GESAMTVERMÖGEN', f"{self.tax_summary['wealth']['total_assets_chf']:.2f}"])
        
        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name='1_Vermögensaufstellung', index=False, header=False)

    def _write_income_sheet(self, writer):
        """Write Einkünfte aus Vermögen sheet"""
        income = self.tax_summary['income']
        foreign_tax = self.tax_summary['foreign_tax']
        
        data = []
        data.append(['EINKÜNFTE AUS VERMÖGEN', 'Betrag (CHF)'])
        data.append(['', ''])
        
        # Dividends by source
        data.append(['Dividenden Schweiz', f"{income['by_source']['swiss']['dividends']:.2f}"])
        data.append(['Dividenden Ausland', f"{income['by_source']['foreign']['dividends']:.2f}"])
        data.append(['Zinsen', f"{income['by_source']['interest']['total']:.2f}"])
        data.append(['', ''])
        
        # Taxes withheld
        data.append(['Quellensteuer', ''])
        data.append(['US Quellensteuer', f"-{foreign_tax['us_withholding_tax']:.2f}"])
        data.append(['Andere Länder', f"-{foreign_tax['other_foreign_tax']:.2f}"])
        data.append(['Gesamt Quellensteuer', f"-{foreign_tax['total_withheld']:.2f}"])
        data.append(['', ''])
        
        # Totals
        data.append(['BRUTTOEINKOMMEN', f"{income['total_income_chf']:.2f}"])
        data.append(['NETTOEINKOMMEN (nach Quellensteuer)', 
                    f"{income['total_income_chf'] - foreign_tax['total_withheld']:.2f}"])
        
        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name='2_Einkünfte', index=False, header=False)

    def _write_capital_gains_sheet(self, writer):
        """Write Kapitalgewinne/Verluste sheet"""
        capital_gains = self.tax_summary['capital_gains']
        
        data = []
        data.append(['KAPITALGEWINNE UND VERLUSTE', 'Betrag (CHF)'])
        data.append(['', ''])
        
        data.append(['Realisierte Gewinne', f"{capital_gains['realized_gains']:.2f}"])
        data.append(['Realisierte Verluste', f"-{capital_gains['realized_losses']:.2f}"])
        data.append(['', ''])
        data.append(['NETTO KAPITALGEWINN/-VERLUST', f"{capital_gains['net_capital_gain_loss']:.2f}"])
        
        # Note about tax treatment
        data.append(['', ''])
        data.append(['Hinweis:', 'Kapitalgewinne aus privatem Vermögen sind in BL steuerfrei'])
        
        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name='3_Kapitalgewinne', index=False, header=False)

    def _write_costs_sheet(self, writer):
        """Write Kosten und Abzüge sheet"""
        costs = self.tax_summary['costs']
        
        data = []
        data.append(['KOSTEN UND ABZÜGE', 'Betrag (CHF)'])
        data.append(['', ''])
        
        data.append(['Transaktionsgebühren', f"{costs['transaction_fees']:.2f}"])
        data.append(['Kontogebühren', f"{costs['account_fees']:.2f}"])
        data.append(['Sonstige Kosten', f"{costs['other_costs']:.2f}"])
        data.append(['', ''])
        data.append(['GESAMT ABZUGSFÄHIGE KOSTEN', f"{costs['total_deductible_costs']:.2f}"])
        
        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name='4_Kosten', index=False, header=False)

    def _write_fx_sheet(self, writer):
        """Write Devisenverluste sheet"""
        fx_result = self.tax_summary['fx_result']
        
        data = []
        data.append(['DEVISENGEWINNE UND -VERLUSTE', 'Betrag (CHF)'])
        data.append(['', ''])
        
        data.append(['Devisengewinne', f"{fx_result['fx_gains']:.2f}"])
        data.append(['Devisenverluste', f"-{fx_result['fx_losses']:.2f}"])
        data.append(['', ''])
        data.append(['NETTO DEVISENERGEBNIS', f"{fx_result['net_fx_result']:.2f}"])
        
        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name='5_Devisen', index=False, header=False)

    def _write_summary_sheet(self, writer):
        """Write summary sheet with tax calculations"""
        wealth = self.tax_summary['wealth']
        income = self.tax_summary['income']
        tax_liability = self.tax_summary['tax_liability']
        
        data = []
        data.append(['STEUERLICHE ZUSAMMENFASSUNG - BASEL-LANDSCHAFT', ''])
        data.append(['Jahr:', str(self.tax_summary['tax_year'])])
        data.append(['Datum:', self.tax_summary['calculation_date']])
        data.append(['', ''])
        
        # Wealth tax
        data.append(['VERMÖGENSSTEUER', ''])
        data.append(['Steuerbares Vermögen (CHF)', f"{wealth['taxable_wealth']:.2f}"])
        data.append(['Steuersatz', f"{wealth['tax_rate']*100:.2f}%"])
        data.append(['Vermögenssteuer (CHF)', f"{wealth['wealth_tax']:.2f}"])
        data.append(['', ''])
        
        # Income tax
        data.append(['EINKOMMENSSTEUER', ''])
        data.append(['Bruttoeinkommen (CHF)', f"{income['gross_income']:.2f}"])
        data.append(['Abzüge (CHF)', f"-{income['deductible_costs']:.2f}"])
        data.append(['Steuerbares Einkommen (CHF)', f"{income['taxable_income']:.2f}"])
        data.append(['Steuersatz', f"{income['tax_rate']*100:.2f}%"])
        data.append(['Einkommenssteuer (CHF)', f"{income['income_tax']:.2f}"])
        data.append(['', ''])
        
        # Tax liability
        data.append(['STEUERSCHULD', ''])
        data.append(['Vermögenssteuer', f"{tax_liability['wealth_tax']:.2f}"])
        data.append(['Einkommenssteuer', f"{tax_liability['income_tax']:.2f}"])
        data.append(['Bruttosteuer', f"{tax_liability['gross_tax_liability']:.2f}"])
        data.append(['Anrechenbare ausländische Quellensteuer', f"-{tax_liability['foreign_tax_credit']:.2f}"])
        data.append(['', ''])
        data.append(['NETTO ZU ZAHLENDE STEUER', f"{tax_liability['net_tax_liability']:.2f}"])
        
        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name='0_ZUSAMMENFASSUNG', index=False, header=False)

    def _apply_excel_formatting(self, writer):
        """Apply formatting to Excel workbook"""
        workbook = writer.book
        
        # Format colors from config
        header_color = self.config.get('excel_format', {}).get('colors', {}).get('header_bg', '4472C4')
        
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            
            # Format title row
            if ws['A1'].value:
                ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
                ws['A1'].fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
                ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

    def generate_pdf_report(self, output_file: str):
        """
        Generate PDF report
        
        Args:
            output_file: Path to output PDF file
        """
        logger.info(f"📄 Generating PDF report: {output_file}")
        
        doc = SimpleDocTemplate(output_file, pagesize=A4,
                               leftMargin=50, rightMargin=50,
                               topMargin=50, bottomMargin=50)
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#4472C4'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#4472C4'),
            spaceAfter=12,
            spaceBefore=12
        )
        
        # Title
        elements.append(Paragraph(f"Steuerreport {self.tax_summary['tax_year']}", title_style))
        elements.append(Paragraph("Canton Basel-Landschaft", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Personal data placeholder
        elements.append(Paragraph("Persönliche Daten", heading_style))
        personal_data = [
            ['Name:', '[Platzhalter]'],
            ['Adresse:', '[Platzhalter]'],
            ['Steuernummer:', '[Platzhalter]'],
            ['Berichtsdatum:', self.tax_summary['calculation_date']]
        ]
        t = Table(personal_data, colWidths=[100, 300])
        t.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
            ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 10),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 20))
        
        # Section 1: Vermögensaufstellung
        elements.append(Paragraph("1. Vermögensaufstellung (Stand 31.12.2025)", heading_style))
        wealth_data = [
            ['Kategorie', 'Wert (CHF)'],
            ['Gesamtvermögen', f"{self.tax_summary['wealth']['total_assets_chf']:.2f}"]
        ]
        t = Table(wealth_data, colWidths=[300, 100])
        t.setStyle(self._get_table_style())
        elements.append(t)
        elements.append(Spacer(1, 15))
        
        # Section 2: Einkünfte
        elements.append(Paragraph("2. Einkünfte aus Vermögen", heading_style))
        income = self.tax_summary['income']
        income_data = [
            ['Kategorie', 'Betrag (CHF)'],
            ['Dividenden Schweiz', f"{income['by_source']['swiss']['dividends']:.2f}"],
            ['Dividenden Ausland', f"{income['by_source']['foreign']['dividends']:.2f}"],
            ['Zinsen', f"{income['by_source']['interest']['total']:.2f}"],
            ['Bruttoeinkommen', f"{income['total_income_chf']:.2f}"]
        ]
        t = Table(income_data, colWidths=[300, 100])
        t.setStyle(self._get_table_style())
        elements.append(t)
        elements.append(Spacer(1, 15))
        
        # Section 3: Capital Gains
        elements.append(Paragraph("3. Kapitalgewinne/Verluste", heading_style))
        cg = self.tax_summary['capital_gains']
        cg_data = [
            ['Kategorie', 'Betrag (CHF)'],
            ['Realisierte Gewinne', f"{cg['realized_gains']:.2f}"],
            ['Realisierte Verluste', f"-{cg['realized_losses']:.2f}"],
            ['Netto', f"{cg['net_capital_gain_loss']:.2f}"]
        ]
        t = Table(cg_data, colWidths=[300, 100])
        t.setStyle(self._get_table_style())
        elements.append(t)
        elements.append(Spacer(1, 15))
        
        # Section 4: Costs
        elements.append(Paragraph("4. Kosten und Abzüge", heading_style))
        costs = self.tax_summary['costs']
        costs_data = [
            ['Kategorie', 'Betrag (CHF)'],
            ['Transaktionsgebühren', f"{costs['transaction_fees']:.2f}"],
            ['Kontogebühren', f"{costs['account_fees']:.2f}"],
            ['Gesamt', f"{costs['total_deductible_costs']:.2f}"]
        ]
        t = Table(costs_data, colWidths=[300, 100])
        t.setStyle(self._get_table_style())
        elements.append(t)
        elements.append(Spacer(1, 15))
        
        # Section 5: FX
        elements.append(Paragraph("5. Devisengewinne/-verluste", heading_style))
        fx = self.tax_summary['fx_result']
        fx_data = [
            ['Kategorie', 'Betrag (CHF)'],
            ['Devisengewinne', f"{fx['fx_gains']:.2f}"],
            ['Devisenverluste', f"-{fx['fx_losses']:.2f}"],
            ['Netto', f"{fx['net_fx_result']:.2f}"]
        ]
        t = Table(fx_data, colWidths=[300, 100])
        t.setStyle(self._get_table_style())
        elements.append(t)
        elements.append(Spacer(1, 20))
        
        # Tax Summary
        elements.append(Paragraph("Steuerliche Zusammenfassung", heading_style))
        tax_lib = self.tax_summary['tax_liability']
        wealth = self.tax_summary['wealth']
        tax_data = [
            ['Steuerart', 'Betrag (CHF)'],
            [f"Vermögenssteuer ({wealth['tax_rate']*100:.2f}%)", f"{tax_lib['wealth_tax']:.2f}"],
            [f"Einkommenssteuer ({income['tax_rate']*100:.2f}%)", f"{tax_lib['income_tax']:.2f}"],
            ['Bruttosteuer', f"{tax_lib['gross_tax_liability']:.2f}"],
            ['Anrechenbare Quellensteuer', f"-{tax_lib['foreign_tax_credit']:.2f}"],
            ['NETTO ZU ZAHLENDE STEUER', f"{tax_lib['net_tax_liability']:.2f}"]
        ]
        t = Table(tax_data, colWidths=[300, 100])
        style = self._get_table_style()
        style.add('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#B4C7E7'))
        style.add('FONT', (0, -1), (-1, -1), 'Helvetica-Bold', 12)
        t.setStyle(style)
        elements.append(t)
        elements.append(Spacer(1, 30))
        
        # Signature placeholder
        elements.append(Paragraph("Unterschrift", heading_style))
        sig_data = [
            ['Datum:', '_' * 30],
            ['Unterschrift:', '_' * 30]
        ]
        t = Table(sig_data, colWidths=[100, 300])
        t.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ]))
        elements.append(t)
        
        # Build PDF
        doc.build(elements)
        
        logger.info(f"✅ PDF report generated: {output_file}")

    def _get_table_style(self) -> TableStyle:
        """Get standard table style for PDF"""
        return TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
        ])

    def generate_json_report(self, output_file: str):
        """
        Generate JSON report
        
        Args:
            output_file: Path to output JSON file
        """
        logger.info(f"📋 Generating JSON report: {output_file}")
        
        # Create comprehensive JSON structure
        json_report = {
            'metadata': {
                'canton': self.tax_summary['canton'],
                'tax_year': self.tax_summary['tax_year'],
                'calculation_date': self.tax_summary['calculation_date'],
                'report_version': '1.0'
            },
            'sections': {
                'section_1_vermoegensaufstellung': {
                    'total_assets_chf': self.tax_summary['wealth']['total_assets_chf'],
                    'wealth_tax_rate': self.tax_summary['wealth']['tax_rate'],
                    'wealth_tax_chf': self.tax_summary['wealth']['wealth_tax']
                },
                'section_2_einkuenfte': {
                    'dividends_swiss': self.tax_summary['income']['by_source']['swiss']['dividends'],
                    'dividends_foreign': self.tax_summary['income']['by_source']['foreign']['dividends'],
                    'interest': self.tax_summary['income']['by_source']['interest']['total'],
                    'total_income_chf': self.tax_summary['income']['total_income_chf'],
                    'foreign_withholding_tax': self.tax_summary['foreign_tax']['total_withheld']
                },
                'section_3_kapitalgewinne': {
                    'realized_gains': self.tax_summary['capital_gains']['realized_gains'],
                    'realized_losses': self.tax_summary['capital_gains']['realized_losses'],
                    'net_capital_gain_loss': self.tax_summary['capital_gains']['net_capital_gain_loss']
                },
                'section_4_kosten': {
                    'transaction_fees': self.tax_summary['costs']['transaction_fees'],
                    'account_fees': self.tax_summary['costs']['account_fees'],
                    'total_deductible_costs': self.tax_summary['costs']['total_deductible_costs']
                },
                'section_5_devisen': {
                    'fx_gains': self.tax_summary['fx_result']['fx_gains'],
                    'fx_losses': self.tax_summary['fx_result']['fx_losses'],
                    'net_fx_result': self.tax_summary['fx_result']['net_fx_result']
                }
            },
            'tax_summary': {
                'wealth_tax': self.tax_summary['tax_liability']['wealth_tax'],
                'income_tax': self.tax_summary['tax_liability']['income_tax'],
                'gross_tax_liability': self.tax_summary['tax_liability']['gross_tax_liability'],
                'foreign_tax_credit': self.tax_summary['tax_liability']['foreign_tax_credit'],
                'net_tax_liability': self.tax_summary['tax_liability']['net_tax_liability']
            },
            'detailed_data': {
                'open_positions': self.parsed_data.get('open_positions', []),
                'transactions_summary': {
                    'count': len(self.parsed_data.get('transactions', [])),
                    'total_proceeds_chf': sum(t.get('proceeds_chf', 0) for t in self.parsed_data.get('transactions', []))
                },
                'dividends_summary': {
                    'count': len([d for d in self.parsed_data.get('dividends', []) if d.get('type') != 'Interest']),
                    'total_amount_chf': self.tax_summary['income']['total_income_chf']
                }
            }
        }
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(json_report, f, indent=2, ensure_ascii=False)
        
        logger.info(f"✅ JSON report generated: {output_file}")


if __name__ == "__main__":
    # Example usage
    logging.basicConfig(level=logging.INFO)
    
    # Example parsed data
    example_data = {
        'transactions': [
            {'type': 'Stocks', 'symbol': 'AAPL', 'proceeds_chf': 1000, 'commission_chf': 5},
            {'type': 'Forex', 'symbol': 'EUR.USD', 'proceeds_chf': 50, 'commission_chf': 0}
        ],
        'dividends': [
            {'symbol': 'AAPL', 'amount_chf': 100, 'type': 'Dividend'},
            {'symbol': 'MSFT', 'amount_chf': 50, 'type': 'Dividend'},
            {'amount_chf': 10, 'type': 'Interest'}
        ],
        'taxes': [
            {'amount_chf': 15, 'country': 'US'}
        ],
        'fees': [
            {'amount_chf': 10, 'type': 'Account fee'}
        ],
        'open_positions': [
            {'symbol': 'AAPL', 'value_chf': 5000, 'currency': 'USD', 'quantity': 10},
            {'symbol': 'MSFT', 'value_chf': 3000, 'currency': 'USD', 'quantity': 5}
        ]
    }
    
    generator = ReportGeneratorBL(example_data)
    reports = generator.generate_all_reports('.')
    
    print("\n✅ Reports generated:")
    for format_type, path in reports.items():
        print(f"  {format_type.upper()}: {path}")
